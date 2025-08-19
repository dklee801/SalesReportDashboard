#!/usr/bin/env python3
"""
매출채권 데이터 자동 복사기 (리팩토링됨)
매출채권_결과.xlsx의 4개 시트를 주간보고서 양식에 자동 복사
"""

import pandas as pd
from pathlib import Path
import logging
from openpyxl import load_workbook
import shutil
from datetime import datetime
import sys

# 리팩토링된 경로 설정
sys.path.append(str(Path(__file__).parent.parent.parent))

# 설정 관리자 import (새 구조)
from modules.utils.config_manager import get_config


class ReceivablesDataCopier:
    """매출채권 데이터 자동 복사기 (리팩토링됨)"""
    
    def __init__(self):
        self.logger = logging.getLogger('ReceivablesDataCopier')
        self.config = get_config()
        
        # 기본 경로 설정 (리팩토링된 경로)
        self.base_dir = Path(__file__).parent.parent.parent.parent
        self.template_file = self.base_dir / "2025년도 주간보고 양식_2.xlsx"
        self.receivables_file = self.config.get_processed_data_dir() / "채권_분석_결과.xlsx"
        
        # 복사할 시트 매핑 (중간파일 → 주간보고서)
        self.sheet_mapping = {
            "매출채권요약": "매출채권요약",
            "90일채권현황": "90일채권현황", 
            "결제기간초과채권현황": "결제기간초과채권현황",
            "결제기간초과채권TOP20": "결제기간초과채권TOP20"
        }
        
        # 각 시트의 복사 설정 (소스 범위 → 타겟 시작 위치)
        self.copy_configs = {
            "매출채권요약": {
                "source_range": "A4:I6",  # 중간파일의 A4부터 읽기 (DND, DNI, 합계 데이터만)
                "target_start": (4, 1)    # 주간보고서의 A4부터 넣기 (row=4, col=1)
            },
            "90일채권현황": {
                "source_range": "A2:D4",  # 중간파일의 A2부터 읽기
                "target_start": (2, 1)    # 주간보고서의 A2부터 넣기
            },
            "결제기간초과채권현황": {
                "source_range": "A2:D4",  # 중간파일의 A2부터 읽기
                "target_start": (2, 1)    # 주간보고서의 A2부터 넣기
            },
            "결제기간초과채권TOP20": {
                "source_range": "A2:G21", # 중간파일의 A2부터 읽기
                "target_start": (2, 1)    # 주간보고서의 A2부터 넣기
            }
        }
    
    def copy_sheet_data_with_offset(self, source_sheet, target_sheet, config):
        """소스 시트에서 타겟 시트로 데이터 복사 (오프셋 적용)"""
        source_range = config["source_range"] 
        target_start_row, target_start_col = config["target_start"]
        
        self.logger.info(f"📋 복사 설정: {source_range} → ({target_start_row}, {target_start_col})")
        
        # 소스 범위 파싱
        start_cell, end_cell = source_range.split(':')
        
        # 소스 데이터 읽기
        source_data = []
        for row in source_sheet[source_range]:
            row_data = []
            for cell in row:
                # None 값 처리
                value = cell.value if cell.value is not None else ""
                row_data.append(value)
            source_data.append(row_data)
        
        # 디버깅: 읽은 소스 데이터 확인
        self.logger.info(f"📋 소스 데이터 읽기 완료: {len(source_data)}행")
        for i, row_data in enumerate(source_data):
            self.logger.info(f"  소스 행 {i}: {row_data[:3]}... (처음 3개 셀만)")
        
        # 타겟에 데이터 쓰기
        for row_idx, row_data in enumerate(source_data):
            target_row = target_start_row + row_idx
            self.logger.info(f"📝 행 {row_idx} 복사: {row_data[:3]}... → 타겟 행 {target_row}")
            
            for col_idx, value in enumerate(row_data):
                target_col = target_start_col + col_idx
                target_sheet.cell(row=target_row, column=target_col, value=value)
        
        self.logger.info(f"✅ 복사 완료: {len(source_data)}행 x {len(source_data[0]) if source_data else 0}열")
    
    def copy_receivables_to_template(self, target_file_path):
        """매출채권 데이터를 지정된 파일에 복사"""
        try:
            self.logger.info("=== 매출채권 데이터 자동 복사 시작 (리팩토링됨) ===")
            
            # 중간파일 존재 확인
            if not self.receivables_file.exists():
                self.logger.error(f"매출채권 중간파일이 없습니다: {self.receivables_file}")
                return False
            
            # 대상 파일 존재 확인
            target_path = Path(target_file_path)
            if not target_path.exists():
                self.logger.error(f"대상 파일이 없습니다: {target_path}")
                return False
            
            # 워크북 열기
            source_wb = load_workbook(str(self.receivables_file))
            target_wb = load_workbook(str(target_path))
            
            # 각 시트 복사
            copied_sheets = 0
            
            for source_sheet_name, target_sheet_name in self.sheet_mapping.items():
                try:
                    # 소스 시트 확인
                    if source_sheet_name not in source_wb.sheetnames:
                        self.logger.warning(f"소스 시트가 없습니다: {source_sheet_name}")
                        continue
                    
                    # 타겟 시트 확인
                    if target_sheet_name not in target_wb.sheetnames:
                        self.logger.warning(f"타겟 시트가 없습니다: {target_sheet_name}")
                        continue
                    
                    # 시트 데이터 복사
                    self.copy_sheet_data_with_offset(
                        source_wb[source_sheet_name],
                        target_wb[target_sheet_name],
                        self.copy_configs[source_sheet_name]
                    )
                    
                    copied_sheets += 1
                    self.logger.info(f"✅ {source_sheet_name} → {target_sheet_name} 복사 완료")
                    
                except Exception as e:
                    self.logger.error(f"❌ {source_sheet_name} 복사 실패: {e}")
                    continue
            
            # 대상 파일 저장
            target_wb.save(str(target_path))
            source_wb.close()
            target_wb.close()
            
            self.logger.info(f"=== 매출채권 데이터 복사 완료: {copied_sheets}개 시트 ===")
            return copied_sheets > 0
            
        except Exception as e:
            self.logger.error(f"매출채권 데이터 복사 중 전체 오류: {e}")
            return False
    
    def copy_to_report(self, report_file_path, create_backup=True):
        """보고서 파일에 매출채권 데이터 복사 (백업 포함)"""
        try:
            report_path = Path(report_file_path)
            
            # 백업 생성
            if create_backup and report_path.exists():
                backup_path = report_path.with_name(
                    f"{report_path.stem}_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}{report_path.suffix}"
                )
                shutil.copy2(report_path, backup_path)
                self.logger.info(f"백업 생성됨: {backup_path.name}")
            
            # 매출채권 데이터 복사
            result = self.copy_receivables_to_template(report_path)
            
            if result:
                print(f"  💾 매출채권 데이터 복사 완료: {report_path.name}")
                self.logger.info(f"매출채권 데이터 복사 성공: {report_path}")
            else:
                print(f"  ❌ 매출채권 데이터 복사 실패: {report_path.name}")
                self.logger.error(f"매출채권 데이터 복사 실패: {report_path}")
                
            return result
            
        except Exception as e:
            error_msg = f"매출채권 데이터 복사 중 오류: {e}"
            print(f"  ❌ {error_msg}")
            self.logger.error(error_msg)
            return False

    def check_receivables_data_availability(self):
        """매출채권 데이터 가용성 확인"""
        try:
            # 중간파일 존재 확인
            if not self.receivables_file.exists():
                self.logger.warning(f"매출채권 중간파일이 없습니다: {self.receivables_file}")
                return False, "매출채권 분석 결과 파일이 없습니다."
            
            # 파일 읽기 시도
            wb = load_workbook(str(self.receivables_file))
            
            # 필요한 시트 존재 확인
            missing_sheets = []
            for sheet_name in self.sheet_mapping.keys():
                if sheet_name not in wb.sheetnames:
                    missing_sheets.append(sheet_name)
            
            wb.close()
            
            if missing_sheets:
                missing_msg = f"필요한 시트가 없습니다: {', '.join(missing_sheets)}"
                self.logger.warning(missing_msg)
                return False, missing_msg
            
            return True, "매출채권 데이터 사용 가능"
            
        except Exception as e:
            error_msg = f"매출채권 데이터 확인 중 오류: {e}"
            self.logger.error(error_msg)
            return False, error_msg


def main():
    """테스트용 메인 함수"""
    try:
        copier = ReceivablesDataCopier()
        
        # 데이터 가용성 확인
        is_available, message = copier.check_receivables_data_availability()
        print(f"매출채권 데이터 상태: {message}")
        
        if not is_available:
            print("매출채권 데이터를 먼저 생성해주세요.")
            return False
        
        # 템플릿 파일로 복사 테스트
        if copier.template_file.exists():
            result = copier.copy_receivables_to_template(copier.template_file)
            if result:
                print("✅ 매출채권 데이터 복사 테스트 성공")
            else:
                print("❌ 매출채권 데이터 복사 테스트 실패")
            return result
        else:
            print(f"템플릿 파일이 없습니다: {copier.template_file}")
            return False
            
    except Exception as e:
        print(f"❌ 테스트 중 오류: {e}")
        return False


if __name__ == "__main__":
    main()
