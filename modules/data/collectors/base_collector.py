"""
데이터 수집 베이스 클래스
BaseDataCollector 추상 클래스와 공통 Selenium 유틸리티 제공
"""

import json
import time
import os
import shutil
import calendar
import re
import logging
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from datetime import datetime, timedelta
import pandas as pd
from pathlib import Path
import sys
from abc import ABC, abstractmethod
from typing import Dict, List, Tuple, Optional

# 경로 설정
sys.path.append(str(Path(__file__).parent.parent))
sys.path.append(str(Path(__file__).parent.parent / "validators"))

# 설정 관리자 import
from modules.utils.config_manager import get_config

# 데이터 검증기 import
try:
    from modules.data.validators.sales_data_validator import SalesDataValidator
    VALIDATOR_AVAILABLE = True
except ImportError:
    print("⚠️ 매출 데이터 검증기를 찾을 수 없습니다. 검증 기능이 비활성화됩니다.")
    VALIDATOR_AVAILABLE = False


class BaseDataCollector(ABC):
    """데이터 수집 베이스 클래스"""
    
    def __init__(self, headless_mode=None):
        self.config = get_config()
        self.selenium_config = self.config.get_selenium_config()
        
        # 헤드리스 모드 강제 비활성화 (항상 브라우저 표시)
        self.selenium_config['headless'] = False
        print(f"👀 브라우저 모드: 항상 표시 (헤드리스 모드 비활성화)")
        
    def js_click(self, driver, element):
        """JavaScript를 이용한 안전한 클릭"""
        driver.execute_script("arguments[0].scrollIntoView(true);", element)
        time.sleep(0.5)
        driver.execute_script("arguments[0].click();", element)

    def launch_driver(self):
        """Chrome 드라이버 실행 - 항상 브라우저 표시 모드"""
        chrome_options = Options()
        
        # 항상 브라우저 표시 모드
        print(f"👀 브라우저 모드: 항상 표시 (헤드리스 모드 비활성화)")
        
        # 다운로드 경로 설정
        paths = self.config.get_paths()
        download_path = str(paths['downloads'])
        prefs = {
            "download.default_directory": download_path,
            "download.prompt_for_download": False,
            "download.directory_upgrade": True
        }
        chrome_options.add_experimental_option("prefs", prefs)
        print(f"📁 다운로드 경로 설정: {download_path}")
        
        if self.selenium_config.get("detach_browser", True):
            chrome_options.add_experimental_option("detach", True)
        
        if self.selenium_config.get("disable_automation_flags", True):
            chrome_options.add_argument("--disable-blink-features=AutomationControlled")
            chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
            chrome_options.add_experimental_option('useAutomationExtension', False)
        
        # 추가 옵션들
        chrome_options.add_argument('--disable-dev-shm-usage')
        chrome_options.add_argument('--disable-background-timer-throttling')
        chrome_options.add_argument('--disable-backgrounding-occluded-windows')
        chrome_options.add_argument('--disable-renderer-backgrounding')
        
        driver = webdriver.Chrome(service=Service(), options=chrome_options)
        
        # 브라우저 최대화 (무조건 실행)
        driver.maximize_window()
        print(f"🖥️ 브라우저 창 최대화 완료")
        
        # 대기 시간 설정
        driver.implicitly_wait(self.selenium_config.get("implicit_wait", 10))
        driver.set_page_load_timeout(self.selenium_config.get("page_load_timeout", 30))
        
        return driver

    def basic_login(self, driver, account):
        """기본 로그인 처리 - 브라우저 표시 모드"""
        wait = WebDriverWait(driver, self.selenium_config.get("implicit_wait", 10))
        company_name = account.get("company_name", "")
        
        print(f"   🔐 {company_name} 로그인 시작...")
        
        # 1. 로그인 페이지로 이동
        driver.get('https://login.ecount.com/Login')
        print(f"   🌐 로그인 페이지 로드 완료")
        
        # 2. 페이지 로딩 대기 (브라우저 모드에 최적화)
        print(f"   ⏳ 페이지 로딩 대기 (3초)...")
        time.sleep(3)
        
        # DOM 준비 상태 확인
        try:
            WebDriverWait(driver, 10).until(
                lambda d: d.execute_script("return document.readyState") == "complete"
            )
            print(f"   ✅ DOM 로딩 완료")
        except:
            print(f"   ⚠️ DOM 대기 타임아웃 - 계속 진행")
        
        try:
            # 3. 요소 대기 타임아웃 설정
            timeout = 10  # 브라우저 모드에 최적화된 타임아웃
            print(f"   🔍 요소 대기 타임아웃: {timeout}초")
            
            # 로그인 필드 입력 (순차적으로)
            print(f"   📝 회사코드 입력: {account['company_code']}")
            com_code_field = wait.until(EC.presence_of_element_located((By.ID, "com_code")))
            com_code_field.clear()
            com_code_field.send_keys(account["company_code"])
            time.sleep(1)  # 브라우저 모드 최적화
            
            print(f"   📝 사용자 ID 입력: {account['user_id']}")
            id_field = WebDriverWait(driver, timeout).until(EC.presence_of_element_located((By.ID, "id")))
            id_field.clear()
            id_field.send_keys(account["user_id"])
            time.sleep(1)
            
            print(f"   📝 비밀번호 입력")
            passwd_field = WebDriverWait(driver, timeout).until(EC.presence_of_element_located((By.ID, "passwd")))
            passwd_field.clear()
            passwd_field.send_keys(account["user_pw"])
            time.sleep(1)  # 브라우저 모드 최적화
            
            # 4. 로그인 버튼 클릭
            print(f"로그인 버튼 클릭")
            login_button = WebDriverWait(driver, timeout).until(EC.element_to_be_clickable((By.ID, "save")))
            
            # 브라우저 모드에서는 일반 클릭 사용
            login_button.click()
            
            # 5. 로그인 처리 대기
            wait_time = 10  # 브라우저 모드 최적화
            print(f"   ⏳ 로그인 처리 대기 ({wait_time}초)...")
            time.sleep(wait_time)
            
            # 6. 로그인 성공 확인
            current_url = driver.current_url
            print(f"   🔍 현재 URL: {current_url}")
            
            if "login.ecount.com" in current_url:
                print(f"   ⚠️ {company_name} 로그인 실패 - 로그인 페이지에 머물러 있음")
                
                # 오류 메시지 확인
                try:
                    error_elements = driver.find_elements(By.CLASS_NAME, "error")
                    if error_elements:
                        for error in error_elements:
                            if error.is_displayed():
                                print(f"   ❌ 오류 메시지: {error.text}")
                except:
                    pass
                
                return False
                
            elif "ecount.com" in current_url:
                print(f"   ✅ {company_name} 로그인 성공!")
                return True
            else:
                print(f"   🤔 {company_name} 로그인 상태 불명 - URL: {current_url}")
                return False
                
        except Exception as e:
            print(f"   ❌ {company_name} 로그인 중 오류: {e}")
            return False
        
        # 7. 추가 안정화 대기
        time.sleep(3)  # 브라우저 모드 최적화

    def wait_for_download(self, company_name: str, target_filename: str, download_timeout: int = None) -> Optional[Path]:
        """개선된 다운로드 대기 및 파일 처리 - Excel 유효성 검증 포함"""
        if download_timeout is None:
            download_timeout = self.config.get_download_timeout()
            
        download_path = self.config.get_downloads_dir()
        print(f"   ⏳ 다운로드 대기 시작 (최대 {download_timeout}초)")
        
        start_time = time.time()
        
        while time.time() - start_time < download_timeout:
            time.sleep(1)
            
            # 개선된 파일 찾기 로직
            xlsx_files = list(download_path.glob("*.xlsx"))
            
            if xlsx_files:
                # 최신 파일 찾기 (생성 시간 기준)
                latest_file = max(xlsx_files, key=lambda x: x.stat().st_ctime)
                
                # 파일 안정성 확인 (크기가 변하지 않는지)
                initial_size = latest_file.stat().st_size
                if initial_size < 1000:  # 1KB 미만이면 아직 다운로드 중
                    continue
                    
                # 안정화 대기 (다운로드 완료 보장을 위해 5초 추가)
                time.sleep(7)
                current_size = latest_file.stat().st_size
                
                if current_size == initial_size and current_size > 1000:
                    # Excel 파일 유효성 검증 (손상된 파일도 반환하여 복구 시도)
                    print(f"   📊 Excel 파일 발견: {latest_file.name} ({current_size:,} bytes)")
                    
                    # 검증 시도하지만 실패해도 파일은 반환 (복구 가능성)
                    is_valid = self.validate_excel_file(latest_file)
                    if is_valid:
                        print(f"   ✅ 유효한 Excel 파일 확인")
                    else:
                        print(f"   ⚠️ Excel 파일 검증 실패 - 복구 시도 예정")
                    
                    return latest_file
                        
        print(f"   ⏰ 다운로드 timeout ({download_timeout}초 초과)")
        return None
    
    def validate_excel_file(self, file_path: Path) -> bool:
        """Excel 파일 유효성 검증 - 회사별 데이터량 차이 반영"""
        try:
            # 기본 파일 크기 확인 (최소 10KB 이상)
            file_size = file_path.stat().st_size
            if file_size < 10 * 1024:  # 10KB 미만은 명백히 비정상
                print(f"   ❌ 파일 크기 부족: {file_size:,} bytes")
                return False
            
            # 회사별 데이터량 차이 반영한 검증
            if file_size < 100 * 1024:  # 100KB 미만
                print(f"   ⚠️ 작은 파일: {file_size:,} bytes (데이터량 적은 회사 가능성)")
            else:
                print(f"   ✅ 파일 크기 정상: {file_size:,} bytes")
            
            # 1차 시도: pandas 기본 읽기
            try:
                df = pd.read_excel(file_path, nrows=5)
                if not df.empty and df.shape[1] > 0:
                    print(f"   ✅ pandas 읽기 성공: {len(df)}행, {df.shape[1]}열")
                    return True
            except Exception as e:
                error_str = str(e).lower()
                if 'stylesheet' in error_str:
                    print(f"   🔧 stylesheet 오류 감지 - 대안 방법 시도")
                else:
                    print(f"   ⚠️ pandas 실패: {str(e)[:50]}...")
            
            # 2차 시도: openpyxl로 데이터만 읽기 (스타일 무시)
            try:
                import openpyxl
                wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
                ws = wb.active
                
                # 첫 5행만 확인
                row_count = 0
                for row in ws.iter_rows(max_row=5, values_only=True):
                    if any(cell is not None for cell in row):
                        row_count += 1
                
                wb.close()
                
                if row_count >= 2:  # 헤더 + 최소 1행 데이터
                    print(f"   ✅ openpyxl 검증 성공: {row_count}행")
                    return True
                else:
                    print(f"   ❌ 유효 데이터 부족: {row_count}행")
                    
            except Exception as e:
                print(f"   ❌ openpyxl 검증 실패: {str(e)[:50]}...")
            
            # 3차 시도: 파일 크기만으로 판단 (대용량 파일은 보통 정상)
            if file_size > 1024 * 1024:  # 1MB 이상이면 OK
                print(f"   ⚠️ 검증 실패하지만 파일 크기로 판단: OK ({file_size:,} bytes)")
                return True
            else:
                print(f"   ❌ 모든 검증 방법 실패")
                return False
                
        except Exception as e:
            print(f"   ❌ 검증 중 오류: {e}")
            return False

    def read_excel_via_csv(self, file_path: Path):
        """Excel→CSV 변환 후 읽기 - stylesheet 오류 완전 우회"""
        
        # 1차 시도: pandas 기본 (빠른 확인)
        try:
            df = pd.read_excel(file_path)
            print(f"   ✅ pandas 직접 읽기 성공: {len(df)}행, {len(df.columns)}열")
            return df
        except Exception as e:
            error_str = str(e).lower()
            if 'stylesheet' in error_str:
                print(f"   🔧 stylesheet 오류 감지 - CSV 변환 모드로 전환")
            else:
                print(f"   ⚠️ pandas 실패: {str(e)[:50]}... - CSV 변환 시도")
        
        # 2차 시도: openpyxl로 데이터만 읽기 (스타일 무시)
        try:
            import openpyxl
            print(f"   🔧 openpyxl 데이터만 읽기 시도...")
            
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb.active
            
            # 모든 데이터 추출
            data = []
            for row in ws.iter_rows(values_only=True):
                if any(cell is not None for cell in row):
                    data.append(row)
            
            wb.close()
            
            if data and len(data) > 1:
                df = pd.DataFrame(data[1:], columns=data[0])
                print(f"   ✅ openpyxl 읽기 성공: {len(df)}행, {len(df.columns)}열")
                return df
            else:
                raise Exception("추출된 데이터 없음")
                
        except Exception as e:
            print(f"   ❌ openpyxl 실패: {str(e)[:50]}...")
        
        # 3차 시도: 최소한의 읽기
        print(f"   ❌ 모든 읽기 방법 실패")
        raise Exception(f"Excel 파일 읽기 실패: {str(e)}")

    def set_headless_mode(self, headless: bool = True):
        """헤드리스 모드 설정"""
        self.selenium_config['headless'] = headless
    
    @abstractmethod
    def get_target_accounts(self) -> List[Dict[str, str]]:
        """대상 계정 목록 반환"""
        pass
    
    @abstractmethod
    def navigate_to_target_page(self, driver, account):
        """대상 페이지로 이동"""
        pass
    
    @abstractmethod
    def set_search_criteria(self, driver, **kwargs):
        """검색 조건 설정"""
        pass
    
    @abstractmethod
    def download_and_save(self, driver, company_name: str, **kwargs) -> bool:
        """다운로드 및 저장"""
        pass
